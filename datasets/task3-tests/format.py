"""
Reformat the 'Dialogue' column of an annotation CSV so each speaker turn
sits on its own line, separated by a blank line, with the role labels
(مريض / مساعد طبي / طبيب / المريض / الطبيب) bolded (when output is .xlsx).

Usage:
    python reformat_dialogue.py input.csv                 # -> input_reformatted.xlsx
    python reformat_dialogue.py input.csv output.csv      # CSV with newlines
    python reformat_dialogue.py input.csv output.xlsx     # XLSX with bold roles
"""

import sys
import re
from pathlib import Path

import pandas as pd

DIALOGUE_COL = "Dialogue"

# Known role labels. Add more here if your dataset uses others.
ROLE_PATTERNS = [
    r"مساعد\s+طبي\s*:",
    r"مريض\s*:",
    r"المريض\s*:",
    r"الطبيب\s*:",
    r"طبيب\s*:",
    r"الدكتور\s*:",
    r"دكتور\s*:",
]
_ROLE_RE = re.compile("(" + "|".join(ROLE_PATTERNS) + ")")


def split_into_turns(text: str):
    """Return a list of (role, content) tuples. role is '' for preamble text."""
    if not isinstance(text, str) or not text.strip():
        return []
    parts = _ROLE_RE.split(text)
    turns = []
    # parts = [preamble, role1, content1, role2, content2, ...]
    if parts[0].strip():
        turns.append(("", parts[0].strip()))
    i = 1
    while i < len(parts):
        role = parts[i].strip()
        content = parts[i + 1].strip() if i + 1 < len(parts) else ""
        turns.append((role, content))
        i += 2
    return turns


def reformat_plain(text: str) -> str:
    """Plain-text reformat: role on same line, blank line between turns."""
    turns = split_into_turns(text)
    if not turns:
        return text
    return "\n\n".join(f"{role} {content}".strip() for role, content in turns)


def build_rich_text(text: str):
    """Build an openpyxl CellRichText with bold role labels."""
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    turns = split_into_turns(text)
    if not turns:
        return text

    bold = InlineFont(b=True)
    plain = InlineFont()
    blocks = []
    for i, (role, content) in enumerate(turns):
        if i > 0:
            blocks.append(TextBlock(plain, "\n\n"))
        if role:
            blocks.append(TextBlock(bold, role + " "))
        if content:
            blocks.append(TextBlock(plain, content))
    return CellRichText(*blocks)


def write_xlsx(df: pd.DataFrame, out_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "annotations"

    # Header
    header_font = Font(bold=True)
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    dialog_col_idx = list(df.columns).index(DIALOGUE_COL) + 1

    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            if j == dialog_col_idx and isinstance(val, str) and val.strip():
                ws.cell(row=i, column=j, value=build_rich_text(val))
            else:
                ws.cell(row=i, column=j, value=val)
            ws.cell(row=i, column=j).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    # Sensible widths
    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        if col == DIALOGUE_COL:
            ws.column_dimensions[letter].width = 80
        elif col in ("Response A", "Response B"):
            ws.column_dimensions[letter].width = 50
        elif col == "Primary Reasoning Objective":
            ws.column_dimensions[letter].width = 45
        else:
            ws.column_dimensions[letter].width = 20

    # Taller rows so the multi-line cell is readable
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 200

    wb.save(out_path)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    in_path = Path(sys.argv[1])
    if not in_path.is_file():
        print(f"File not found: {in_path}")
        sys.exit(1)

    if len(sys.argv) > 2:
        out_path = Path(sys.argv[2])
    else:
        out_path = in_path.with_name(in_path.stem + "_reformatted.xlsx")

    df = pd.read_csv(in_path, dtype=str, keep_default_na=False)

    # Clean up headers: strip whitespace and drop empty "Unnamed: N" phantom columns
    df.columns = [str(c).strip() for c in df.columns]
    df = df.loc[:, ~df.columns.str.match(r"^Unnamed:\s*\d+$")]
    # Also drop any columns that are entirely empty strings
    df = df.loc[:, ~(df.apply(lambda col: col.astype(str).str.strip().eq("")).all())]

    if DIALOGUE_COL not in df.columns:
        print(f"Column '{DIALOGUE_COL}' not found. Columns in file: {list(df.columns)}")
        sys.exit(1)

    # Quick stats on turn-splitting
    n_total = len(df)
    turn_counts = df[DIALOGUE_COL].map(lambda t: len(split_into_turns(t)))
    n_unsplit = int((turn_counts <= 1).sum())
    print(f"Rows: {n_total}")
    print(f"Avg turns per dialogue: {turn_counts.mean():.1f}")
    print(f"Rows where no role markers were found: {n_unsplit}")
    if n_unsplit > 0:
        print("  -> those rows were left unchanged. Check role patterns at top of script.")

    if out_path.suffix.lower() in (".xlsx", ".xlsm"):
        write_xlsx(df, out_path)
    else:
        df[DIALOGUE_COL] = df[DIALOGUE_COL].map(reformat_plain)
        df.to_csv(out_path, index=False, encoding="utf-8-sig")

    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()