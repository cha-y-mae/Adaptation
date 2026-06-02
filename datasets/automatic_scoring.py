"""
Auto-label LLM predictions that exactly match the ground-truth reference.

Input CSV columns (as exported from the Google Sheet):
    ID | QUESTION | GROUND TRUTH | MODEL | PREDICTION | LABEL

The first three columns come from merged cells: only the first row of each
question group has the value, the remaining model rows are blank. This script
forward-fills them so every row has its reference.

For each row, PREDICTION is compared to GROUND TRUTH using a NORMALIZED
Arabic comparison:
    - strip leading/trailing whitespace
    - collapse runs of whitespace to a single space
    - remove tashkeel (diacritics)
    - remove tatweel (kashida)
    - unify alef variants (أ إ آ ٱ -> ا)
    - unify alef maksura  (ى -> ي)
    - unify ta marbuta    (ة -> ه)  [toggle NORMALIZE_TA_MARBUTA below]

If the normalized prediction equals the normalized ground truth, LABEL is
set to "correct". Otherwise, the existing LABEL value is preserved (so any
human labels already entered are NOT overwritten).

Usage:
    python auto_label_exact_match.py input.csv                  # writes input_labeled.csv
    python auto_label_exact_match.py input.csv output.csv       # custom output path
"""

import sys
import re
import unicodedata
from pathlib import Path

import pandas as pd

# ---------- CONFIG ----------
MERGED_COLS = ["ID", "QUESTION", "GROUND TRUTH"]
MODEL_COL = "MODEL"
PREDICTION_COL = "PREDICTION"
LABEL_COL = "LABEL"
CORRECT_LABEL = "correct"

# If True, treat ة and ه as equivalent. Some annotators want this, some don't.
NORMALIZE_TA_MARBUTA = False
# ----------------------------


# Arabic diacritics (tashkeel) and tatweel
_TASHKEEL = re.compile(r"[\u0617-\u061A\u064B-\u0652\u0670\u06D6-\u06ED]")
_TATWEEL = "\u0640"
_WHITESPACE = re.compile(r"\s+")


def normalize_arabic(text: str) -> str:
    """Normalize Arabic text for exact-match comparison."""
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)

    # Unicode NFC normalization (canonical composition)
    text = unicodedata.normalize("NFC", text)

    # Remove tashkeel (diacritics) and tatweel
    text = _TASHKEEL.sub("", text)
    text = text.replace(_TATWEEL, "")

    # Unify alef variants
    for ch in ("\u0623", "\u0625", "\u0622", "\u0671"):  # أ إ آ ٱ
        text = text.replace(ch, "\u0627")  # ا

    # Unify alef maksura -> ya
    text = text.replace("\u0649", "\u064a")  # ى -> ي

    # Optional: ta marbuta -> ha
    if NORMALIZE_TA_MARBUTA:
        text = text.replace("\u0629", "\u0647")  # ة -> ه

    # Collapse whitespace and trim
    text = _WHITESPACE.sub(" ", text).strip()

    return text


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    in_path = Path(sys.argv[1])
    if not in_path.is_file():
        print(f"File not found: {in_path}")
        sys.exit(1)

    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else in_path.with_name(
        in_path.stem + "_labeled.csv"
    )

    # Read as strings to avoid any type coercion on Arabic content
    df = pd.read_csv(in_path, dtype=str, keep_default_na=False)

    # Sanity-check required columns
    missing = [c for c in MERGED_COLS + [MODEL_COL, PREDICTION_COL] if c not in df.columns]
    if missing:
        print(f"Missing required columns: {missing}")
        print(f"Found columns: {list(df.columns)}")
        sys.exit(1)

    # Ensure LABEL column exists
    if LABEL_COL not in df.columns:
        df[LABEL_COL] = ""

    # Forward-fill merged cells (ID / QUESTION / GROUND TRUTH):
    # blanks inherit the last non-blank value above them.
    for col in MERGED_COLS:
        df[col] = df[col].replace("", pd.NA).ffill().fillna("")

    # Compute normalized columns (kept in memory; not written to output)
    gt_norm = df["GROUND TRUTH"].map(normalize_arabic)
    pred_norm = df[PREDICTION_COL].map(normalize_arabic)

    # Identify matches: both non-empty AND normalized-equal
    both_nonempty = (gt_norm != "") & (pred_norm != "")
    matches = both_nonempty & (gt_norm == pred_norm)

    # Only overwrite LABEL where it's currently empty (protect human labels)
    empty_label = df[LABEL_COL].fillna("").str.strip() == ""
    to_fill = matches & empty_label

    df.loc[to_fill, LABEL_COL] = CORRECT_LABEL

    # ---------- Stats ----------
    total_rows = len(df)
    total_matches = int(matches.sum())
    newly_labeled = int(to_fill.sum())
    already_labeled_matches = int((matches & ~empty_label).sum())
    n_questions = df["ID"].nunique()

    print(f"\n=== Summary ===")
    print(f"Total rows:             {total_rows}")
    print(f"Unique questions:       {n_questions}")
    print(f"Exact matches found:    {total_matches}")
    print(f"  -> newly labeled:     {newly_labeled}")
    print(f"  -> already had label: {already_labeled_matches}")
    print(f"Rows left for humans:   {total_rows - int((df[LABEL_COL].str.strip() != '').sum())}")

    # Per-model breakdown
    print(f"\n=== Exact matches by model ===")
    per_model = (
        df.assign(_match=matches)
          .groupby(MODEL_COL)["_match"]
          .agg(["sum", "count"])
          .rename(columns={"sum": "matches", "count": "total"})
          .sort_values("matches", ascending=False)
    )
    per_model["pct"] = (per_model["matches"] / per_model["total"] * 100).round(1)
    print(per_model.to_string())

    # ---------- Restore the "merged look" ----------
    # Matching is done; now blank out duplicate values in the merged columns so
    # only the FIRST row of each question group carries the ID / QUESTION /
    # GROUND TRUTH text. This mirrors how the original sheet exported.
    first_in_group = df["ID"] != df["ID"].shift()
    for col in MERGED_COLS:
        df.loc[~first_in_group, col] = ""

    # ---------- Write output ----------
    if out_path.suffix.lower() in (".xlsx", ".xlsm"):
        # Write as Excel with REAL merged cells so it opens cleanly in Sheets/Excel.
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "annotations"

        # Header
        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=1, column=j, value=col)

        # Data: write the FULL (forward-filled) values so no row is blank in the
        # underlying data, then merge ranges for each question group.
        full = df.copy()
        # Re-fill so the cells we just blanked are filled again for writing
        for col in MERGED_COLS:
            full[col] = full[col].replace("", pd.NA).ffill().fillna("")

        for i, row in enumerate(full.itertuples(index=False), start=2):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        # Merge ID / QUESTION / GROUND TRUTH ranges per question group
        group_starts = full.index[first_in_group.values].tolist() + [len(full)]
        col_letters = {c: get_column_letter(list(df.columns).index(c) + 1) for c in MERGED_COLS}
        for a, b in zip(group_starts[:-1], group_starts[1:]):
            if b - a > 1:
                for c, letter in col_letters.items():
                    ws.merge_cells(f"{letter}{a+2}:{letter}{b+1}")

        wb.save(out_path)
    else:
        df.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()